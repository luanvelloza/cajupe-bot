from models.Bot import Bot
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import load_workbook, Workbook
import time
import os
from openpyxl import Workbook, load_workbook

class Apportion_accounts_payable(Bot):
    """Class responsible for apportioning the payment."""

    def _open_filter(self) -> None:
        """Click the button to open the field filter."""
        bnt = self.driver.find_element(By.XPATH, "//*[@id=\"btn--CriarItemSuperSet\"]")
        bnt.click()

    def _write_name(self, name: str) -> None:
        """Write customer's name in the filter"""
        time.sleep(0.5)
        input = self.driver.find_element(By.XPATH, "//div[@class='v-card v-sheet theme--light']//div[@class='v-text-field__slot']//input")
        input.click()
        input.send_keys(Keys.CONTROL, 'a')
        time.sleep(0.5)
        input.send_keys(name)

    def _check_Element(self, xpath: str, time: int) -> bool:
        """
            Check if the element exist and load. If don't load, returns False
        
            Input:
                - xpath - str - (element's xpath)
                - time - int - (How mach time need to load)

            Output:
                - True
                - False
        """
        
        try:
            wait = WebDriverWait(self.driver, timeout=time)
            return wait.until(lambda d: self.driver.find_element(By.XPATH, xpath).is_displayed())
        except:
            return False   
            
    def _click_name_element(self) -> None:
        """Click on the single element on the screen."""
        time.sleep(1)
        bnt = self.driver.find_element(By.XPATH, "//tbody//tr[@class='ATIVO']")
        bnt.click()

    def _write_value_element(self, value: str) -> None:
        """Write the value of the element."""
        time.sleep(1)
        input = self.driver.find_element(By.XPATH, "//div[@class='col-md-6 col-12']//input[@maxlength='14']")
        input.click()
        time.sleep(0.5)
        input.send_keys(value)

    def _click_confirm(self) -> None:
        """Click the confirm button"""
        time.sleep(0.5)
        bnt = self.driver.find_element(By.XPATH, "//button[@id='btn--adicionar']") 
        bnt.click()

    def _cancel_btn(self) -> None:
        """If there is an error, click the cancel button."""
        time.sleep(0.5)
        element = self.driver.find_element(By.XPATH, "//div[@class='v-card v-sheet theme--light']//div[@class='v-card__actions']//button[@style='height: 39px;']")
        element.click()

    def _take_excel_file(self, url) -> dict:
        """
            Take a excel file and return a list of yours elements

            input:
                - file excel' URL

            output:
                - list with id and value.
        """

        pw = load_workbook(url)
        pw_ac = pw.active

        list_temp = []

        for item in pw_ac['A'][1:]:
            id = str(item.value)
            line = item.row
            value = pw_ac[f'B{line}'].value

            if not value is None:
                value = '%.2f' % float(value)
                
                list_temp.append({
                    'id': id,
                    'value': str(value)
                })
        
        return list_temp
    
    def _make_a_table(self, url: str, list: list, name: str) -> None:
        """
            Create a table with list insertion

            Input:
                - url - Need put the "/" in the end of url
                - list
                - name (Table's name - Need put the file extension)

            Output
                - Create a Table into the Url passed
        """

        wb = Workbook()
        ws = wb.active

        ws['A1'].value = 'id'
        ws['B1'].value = 'value'
        row = 2
        
        for item in list:
            id = item['id']
            value = item['value']

            ws[f'A{row}'].value = str(id).upper()
            ws[f'B{row}'].value = value

            row += 1

        wb.save(f'{url}{name}')

    def run_bot(self) -> None:
        self.init_driver()
        list_temp = self._take_excel_file(f'{os.getcwd()}/excel_tables/tables-with-the-apportionment-values/apportionment-values.xlsx')
        self.open_site()
        error_list = []
        list_length = len(list_temp)
        i = 1

        for item in list_temp:

            print(f'{i}/{list_length} | Nome do Cliente: {item["id"]} | Valor: {item["value"]} | Numero de erros: {len(error_list)}')
            i += 1
            self._open_filter()
            self._write_name(item["id"])
            check = self._check_Element("//tbody//tr[@class='ATIVO']", 2)
            
            if check:
                self._click_name_element()
                self._write_value_element(item["value"])
                self._click_confirm()
            else:
                self._cancel_btn()
                error_list.append({"id":item["id"], "value":item["value"]})

        self._make_a_table(f'{os.getcwd()}/excel_tables/tables-with-the-apportionment-values/', error_list, 'to-correct.xlsx')

        input("Digite qualquer tecla para finalizar: ")
        self.driver.quit()
